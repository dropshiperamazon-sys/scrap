"""Exercises main.run()'s orchestration (batching, dedup/merge, resume,
--limit, immediate Excel writes) end-to-end, with discovery/enrichment/
the browser mocked out -- this sandbox can't reach google.com to run the
real Playwright-driven pieces, but the wiring between main.py, database,
and export can be verified fully offline.
"""

import argparse

from openpyxl import load_workbook

import main
from scraper.discovery import DiscoveryDiagnostics, RawBusiness
from scraper.enrichment import EnrichedLead
from utils.config_loader import build_app_config, load_config


class DummyBrowser:
    def __init__(self, headless, logger):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def safe_new_page(self):
        return None

    def close_page(self, page):
        pass


def make_config(tmp_path, **overrides):
    raw = load_config(str(tmp_path / "missing.yaml"))
    defaults = {
        "keywords": ["furniture store"],
        "locations": ["Chicago, IL"],
        "database_file": str(tmp_path / "leads.db"),
        "output_file": str(tmp_path / "leads.xlsx"),
        "log_file": str(tmp_path / "app.log"),
        "batch_size": 2,
        "min_delay_seconds": 0,
        "max_delay_seconds": 0,
    }
    defaults.update(overrides)
    return build_app_config(raw, defaults)


def make_args(**overrides):
    namespace = argparse.Namespace(
        test=False, resume=False, headless=False, limit=None, keyword=None, location=None, config="config.yaml"
    )
    for key, value in overrides.items():
        setattr(namespace, key, value)
    return namespace


def fake_enrich_business(browser, raw, logger, find_email=True, find_pinterest=True):
    domain = raw.website.replace("https://", "") if raw.website else ""
    return EnrichedLead(
        company_name=raw.company_name,
        website=raw.website,
        phone=raw.phone,
        email=f"info@{domain}" if domain else "",
        pinterest="",
        website_reachable=bool(raw.website),
        status="COMPLETE",
    )


def _patch_common(monkeypatch, discover_fn):
    monkeypatch.setattr(main, "discover_businesses", discover_fn)
    monkeypatch.setattr(main, "enrich_business", fake_enrich_business)
    monkeypatch.setattr(main, "BrowserManager", DummyBrowser)
    monkeypatch.setattr(main, "wait_between_batches", lambda *a, **k: None)


def test_run_processes_businesses_and_writes_excel_immediately(tmp_path, monkeypatch):
    businesses = [
        RawBusiness(company_name="Example Furniture", website="https://example-real.com", phone="+1 111-111-1111"),
        RawBusiness(company_name="ABC Home", website="https://abc-real.com", phone="+1 222-222-2222"),
    ]
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=2)
    call_count = {"discover": 0}

    def fake_discover(browser, keyword, location, max_results, logger):
        call_count["discover"] += 1
        return businesses, diagnostics

    _patch_common(monkeypatch, fake_discover)
    config = make_config(tmp_path)
    stats = main.RunStats()

    main.run(config, make_args(), stats)

    assert call_count["discover"] == 1
    assert stats.new_leads == 2
    assert stats.emails_found == 2
    assert stats.duplicates_skipped == 0

    workbook = load_workbook(config.output_file)
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][1] == "Example Furniture"
    assert rows[1][1] == "ABC Home"


def test_resume_skips_already_completed_search(tmp_path, monkeypatch):
    businesses = [RawBusiness(company_name="Example Furniture", website="https://example-real.com", phone="+1 111-111-1111")]
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=1)
    call_count = {"discover": 0}

    def fake_discover(browser, keyword, location, max_results, logger):
        call_count["discover"] += 1
        return businesses, diagnostics

    _patch_common(monkeypatch, fake_discover)
    config = make_config(tmp_path)

    main.run(config, make_args(), main.RunStats())
    assert call_count["discover"] == 1

    main.run(config, make_args(resume=True), main.RunStats())
    assert call_count["discover"] == 1  # unchanged: the completed search was skipped entirely


def test_same_business_across_keywords_is_deduped_not_duplicated(tmp_path, monkeypatch):
    business = RawBusiness(company_name="Example Furniture", website="https://example-real.com", phone="+1 111-111-1111")
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=1)

    def fake_discover(browser, keyword, location, max_results, logger):
        return [business], diagnostics

    _patch_common(monkeypatch, fake_discover)
    config = make_config(tmp_path, keywords=["furniture store", "furniture manufacturer"])
    stats = main.RunStats()

    main.run(config, make_args(), stats)

    assert stats.new_leads == 1
    assert stats.duplicates_skipped == 1  # found again under the second keyword, not re-added

    workbook = load_workbook(config.output_file)
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1  # no duplicate row in the final Excel


def test_limit_stops_after_n_new_leads_not_duplicates(tmp_path, monkeypatch):
    businesses = [
        RawBusiness(company_name=f"Company {i}", website=f"https://company{i}.com", phone=f"+1 000-000-{i:04d}")
        for i in range(5)
    ]
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=5)

    def fake_discover(browser, keyword, location, max_results, logger):
        return businesses, diagnostics

    _patch_common(monkeypatch, fake_discover)
    config = make_config(tmp_path, batch_size=2)
    stats = main.RunStats()

    main.run(config, make_args(limit=2), stats)

    assert stats.new_leads == 2


def test_keywords_and_locations_processed_count_distinct_values(tmp_path, monkeypatch):
    business = RawBusiness(company_name="Example Furniture", website="https://example-real.com", phone="+1 111-111-1111")
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=1)

    def fake_discover(browser, keyword, location, max_results, logger):
        return [business], diagnostics

    _patch_common(monkeypatch, fake_discover)
    # 2 keywords x 2 locations = 4 combinations, but only 2 distinct keywords/locations.
    config = make_config(tmp_path, keywords=["a", "b"], locations=["X", "Y"])
    stats = main.RunStats()

    main.run(config, make_args(), stats)

    assert stats.keywords_processed == 2
    assert stats.locations_processed == 2


def test_require_website_filters_businesses_without_one(tmp_path, monkeypatch):
    businesses = [
        RawBusiness(company_name="Has Website", website="https://has-website.com", phone="+1 111-111-1111"),
        RawBusiness(company_name="No Website", website="", phone="+1 222-222-2222"),
    ]
    diagnostics = DiscoveryDiagnostics(feed_found=True, card_count=2)

    def fake_discover(browser, keyword, location, max_results, logger):
        return businesses, diagnostics

    _patch_common(monkeypatch, fake_discover)
    config = make_config(tmp_path, require_website=True)
    stats = main.RunStats()

    main.run(config, make_args(), stats)

    assert stats.new_leads == 1
    workbook = load_workbook(config.output_file)
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][1] == "Has Website"
