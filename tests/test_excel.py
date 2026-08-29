from openpyxl import load_workbook

from export.excel import COLUMNS, write_excel


def test_write_excel_creates_file_with_exact_columns(tmp_path):
    rows = [
        {"email": "info@example.com", "company_name": "Example Furniture", "website": "https://example.com", "phone": "+1 212-355-9100", "pinterest": "https://pinterest.com/example"},
        {"email": "", "company_name": "ABC Home", "website": "https://abc.com", "phone": "", "pinterest": ""},
    ]
    out_path = tmp_path / "leads.xlsx"
    write_excel(rows, str(out_path))

    workbook = load_workbook(str(out_path))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == COLUMNS

    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert data_rows[0] == ("info@example.com", "Example Furniture", "https://example.com", "+1 212-355-9100", "https://pinterest.com/example")
    # openpyxl round-trips an empty string as a blank cell (None) -- that's
    # correct Excel behavior, not a bug in the exporter.
    assert data_rows[1] == (None, "ABC Home", "https://abc.com", None, None)


def test_write_excel_overwrites_existing_file_atomically(tmp_path):
    out_path = tmp_path / "leads.xlsx"
    write_excel([{"email": "a@example.com", "company_name": "A", "website": "", "phone": "", "pinterest": ""}], str(out_path))
    write_excel([{"email": "b@example.com", "company_name": "B", "website": "", "phone": "", "pinterest": ""}], str(out_path))

    workbook = load_workbook(str(out_path))
    sheet = workbook.active
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 1
    assert data_rows[0][1] == "B"
